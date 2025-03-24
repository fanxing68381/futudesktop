import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 源文件路径和工作表名称
source_file = r'E:\图片\每笔日交易量价模板.xlsx'
source_sheet = '港股每笔日交易量价模板新'

# 读取源数据（E到AA列）
source_wb = load_workbook(source_file)
source_ws = source_wb[source_sheet]

# 目标目录路径
target_dir = r'C:\Users\Administrator\Desktop\每日统计改名'

# 遍历目标目录下的所有.xlsx文件
for filename in os.listdir(target_dir):
    if filename.endswith('.xlsx'):
        target_path = os.path.join(target_dir, filename)

        # 打开目标工作簿
        target_wb = load_workbook(target_path)

        # 获取目标工作簿的第一个工作表
        target_ws = target_wb.active

        # 定义要复制的列范围（E列到AA列，列索引5到27）
        start_col = 4
        end_col = 27

        # 逐行逐列复制数据和填充颜色
        for row in range(1, source_ws.max_row + 1):
            for col in range(start_col, end_col + 1):
                source_cell = source_ws.cell(row=row, column=col)
                target_cell = target_ws.cell(row=row, column=col)

                # 复制单元格的值
                target_cell.value = source_cell.value

                # 复制单元格的填充颜色
                if source_cell.fill.start_color.index != '00000000':
                    fill = PatternFill(
                        fill_type=source_cell.fill.fill_type,
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color
                    )
                    target_cell.fill = fill

        # 保存并关闭目标文件
        target_wb.save(target_path)
        target_wb.close()
        print(f"已更新并保存文件: {target_path}")

# 关闭源工作簿
source_wb.close()
print("操作完成！所有文件已更新并保存。")