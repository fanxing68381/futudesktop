import os
from openpyxl import load_workbook
from datetime import time, datetime

def process_excel_files(directory):
    # 用于存储已处理的文件名
    processed_files = []
    # 文件计数器
    processed_count = 0

    # 遍历目录下的所有文件
    for filename in os.listdir(directory):
        if not filename.endswith('.xlsx'):  # 只处理.xlsx文件
            continue

        filepath = os.path.join(directory, filename)

        try:
            # 使用openpyxl加载Excel文件，保留所有格式
            wb = load_workbook(filepath)
            ws = wb.active

            # 如果工作表为空，跳过处理
            if ws.max_row <= 1:  # 只有表头或空表
                print(f"文件 {filename} 无数据，跳过处理")
                continue

            # 标记要删除的行号
            rows_to_delete = []
            first_15_found = False

            # 从第2行开始处理（跳过表头）
            for row in range(2, ws.max_row + 1):
                cell_value = ws[f'A{row}'].value

                # 处理空值或非时间格式
                if not cell_value:
                    continue

                # 如果是datetime对象，提取time
                if isinstance(cell_value, datetime):
                    t_time = cell_value.time()
                else:
                    # 尝试将字符串解析为时间
                    try:
                        t_time = datetime.strptime(str(cell_value), '%H:%M:%S').time()
                    except ValueError:
                        continue  # 无效时间格式，保留该行

                # 条件1：时间 > 15:00:00 的行标记为删除
                if t_time > time(15, 0, 0):
                    rows_to_delete.append(row)
                # 条件2：处理15:00:00，仅保留第一个，除非C列值不为0
                elif t_time == time(15, 0, 0):
                    # 检查C列值
                    c_value = ws[f'C{row}'].value
                    # 如果C列值不为0（且不是None），保留该行
                    if c_value is not None and c_value != 0:
                        continue  # 保留该行，不标记为删除
                    # 否则按原逻辑处理：仅保留第一个15:00:00
                    if first_15_found:
                        rows_to_delete.append(row)
                    else:
                        first_15_found = True

            # 如果有需要删除的行，从后向前删除以避免行号错乱
            if rows_to_delete:
                for row in sorted(rows_to_delete, reverse=True):
                    ws.delete_rows(row, 1)

            # 保存修改后的文件
            wb.save(filepath)
            processed_files.append(filename)  # 记录处理成功的文件名
            processed_count += 1  # 计数器加1
            print(f"已成功处理文件: {filename}")

        except Exception as e:
            # 如果处理失败，保留原始文件并打印错误信息
            print(f"处理文件 {filename} 失败: {str(e)}")
            continue

    # 统计并输出处理结果
    print("\n=== 处理统计 ===")
    print(f"共处理了 {processed_count} 个文件")
    if processed_count > 0:
        print("处理的文件列表如下：")
        for file in processed_files:
            print(f"- {file}")
    else:
        print("没有文件被成功处理")

# 指定目录路径
directory = r"C:\Users\Administrator\Desktop\每日统计改名"
process_excel_files(directory)