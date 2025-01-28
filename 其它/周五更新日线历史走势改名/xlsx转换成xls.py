import os
import pandas as pd
from openpyxl import load_workbook
import xlwt

# 源目录
source_dir = r'C:\Users\Administrator\Desktop\日线历史走势改名'
# 目标目录
target_dir = r'E:\图片\日线资金历史走势'

# 文件名列表
file_names = [
    '国电南瑞日线历史走势.xls', '华电国际日线历史走势.xls', '海尔智家日线历史走势.xls',
    '乐普医疗日线历史走势.xls', '华强科技日线历史走势.xls', '山东黄金日线历史走势.xls',
    '顺丰控股日线历史走势.xls', '中国海油日线历史走势.xls', '三一重工日线历史走势.xls',
    '潍柴动力日线历史走势.xls', '新希望日线历史走势.xls', '银河磁体日线历史走势.xls',
    '中国平安日线历史走势.xls', '中粮糖业日线历史走势.xls', '中芯国际日线历史走势.xls',
    '中油资本日线历史走势.xls', '华泰证券日线历史走势.xls', '长城汽车日线历史走势.xls',
    '马钢股份日线历史走势.xls', '中航电子日线历史走势.xls', '中金公司日线历史走势.xls',
    '中信证券日线历史走势.xls', '中国船舶日线历史走势.xls', '中国移动日线历史走势.xls',
    '宝钢股份日线历史走势.xls', '紫金矿业日线历史走势.xls', '比亚迪日线历史走势.xls'
]

for file_name in file_names:
    # 构建源文件路径
    source_file = os.path.join(source_dir, file_name)
    # 构建目标文件路径，将源文件扩展名 .xls 替换为 .xlsx
    target_xlsx_file = os.path.join(target_dir, file_name.replace('.xls', '.xlsx'))
    # 构建转换后的目标文件路径，使用 .xls 扩展名
    target_xls_file = os.path.join(target_dir, file_name)

    # 检查源文件和目标文件是否存在
    if os.path.exists(source_file) and os.path.exists(target_xlsx_file):
        try:
            # 读取源文件数据（跳过表头）
            source_df = pd.read_excel(source_file, skiprows=1)

            # 加载目标文件
            wb = load_workbook(target_xlsx_file)
            ws = wb.active

            # 找到目标文件的最后一行
            last_row = ws.max_row + 1

            # 将源文件的数据逐行写入目标文件
            for row in source_df.values.tolist():
                ws.append(row)

            # 保存修改后的目标文件
            wb.save(target_xlsx_file)
            print(f"成功处理 {file_name} 并保存为 .xlsx 文件")

            # 创建新的 xlwt 工作簿
            new_wb = xlwt.Workbook()
            new_ws = new_wb.add_sheet('Sheet1')

            # 遍历 .xlsx 文件的每一行和每一列，将数据写入 .xls 文件
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                for col_idx, value in enumerate(row):
                    new_ws.write(row_idx, col_idx, value)

            # 保存为 .xls 文件
            new_wb.save(target_xls_file)
            print(f"成功将 {target_xlsx_file} 转换为 {target_xls_file}")
        except Exception as e:
            print(f"处理 {file_name} 时出错: {e}")
    else:
        print(f"源文件或目标文件 {file_name} 不存在")

