import pandas as pd
import akshare as ak
import os
import copy
from datetime import datetime, timedelta
from openpyxl import load_workbook

TARGET_FILE = r'E:\图片\每日大盘涨跌——据成交量——趋势判断.xlsx'


def validate_input_date(date_str):
    """日期验证增强版"""
    try:
        input_date = datetime.strptime(date_str, "%Y%m%d")
        if input_date > datetime.now():
            print("警告：不能输入未来日期")
            return False
        if input_date.weekday() >= 5:
            print("注意：输入日期为周末，数据可能无效")
        return True
    except ValueError:
        print("错误：日期格式无效，请使用YYYYMMDD格式")
        return False


def fetch_corrected_data(symbol, date_str):
    """获取并校验交易数据"""
    try:
        trade_date = datetime.strptime(date_str, "%Y%m%d").strftime("%Y%m%d")
        max_retries = 3
        for attempt in range(max_retries):
            try:
                df = ak.index_zh_a_hist_min_em(
                    symbol=symbol,
                    period='15',
                    start_date=trade_date,
                    end_date=trade_date
                )
                break
            except Exception as e:
                if attempt == max_retries - 1:
                    raise
                print(f"请求失败，正在重试({attempt + 1}/{max_retries})...")
                continue

        if df.empty:
            raise ValueError("当日无交易数据")

        required_cols = ['成交量', '成交额']
        missing = set(required_cols) - set(df.columns)
        if missing:
            raise ValueError(f"数据缺失关键列: {missing}")

        if (df['成交量'] <= 0).any() or (df['成交额'] <= 0).any():
            raise ValueError("存在无效的成交数据")

        return df[['时间', '成交量', '成交额']]
    except Exception as e:
        raise RuntimeError(f"数据获取失败: {str(e)}")


def update_workbook_first_sheet(wb, data):
    """精准更新第一个工作表的 H3:I18"""
    try:
        ws = wb.worksheets[0]
        if ws.max_row < 18 or ws.max_column < 9:
            raise ValueError("工作表格式不符合要求")

        if len(data) != 16:
            raise ValueError("需要恰好16条数据填充H3:I18区域")

        for idx in range(16):
            row_num = idx + 3
            amount_billion = data.iloc[idx]['成交额'] / 1e8
            volume_wan = data.iloc[idx]['成交量'] / 1e4
            ws[f'H{row_num}'] = round(amount_billion, 2)
            ws[f'I{row_num}'] = round(volume_wan, 2)
            ws[f'H{row_num}'].number_format = '#,##0.00_ '
            ws[f'I{row_num}'].number_format = '#,##0.00_ '
        return wb
    except Exception as e:
        raise RuntimeError(f"文件更新失败: {str(e)}")


def get_previous_trading_day(input_date):
    """计算前一个交易日（跳过周末）"""
    dt = datetime.strptime(input_date, "%Y%m%d")
    delta = 1
    while True:
        prev_day = dt - timedelta(days=delta)
        if prev_day.weekday() < 5:
            return prev_day
        delta += 1


def get_next_trading_day(input_date):
    """计算下一个交易日（跳过周末）"""
    dt = datetime.strptime(input_date, "%Y%m%d")
    delta = 1
    while True:
        next_day = dt + timedelta(days=delta)
        if next_day.weekday() < 5:
            return next_day
        delta += 1


def copy_cell_style(src_cell, dest_cell):
    """安全样式复制"""
    attrs = ['font', 'border', 'fill', 'alignment', 'number_format']
    for attr in attrs:
        setattr(dest_cell, attr, copy.copy(getattr(src_cell, attr)) if getattr(src_cell, attr) else None)



def clone_worksheet(wb, source_ws):
    """克隆工作表并设置 B23:E23 为前一交易日的公式引用"""
    try:
        # 解析源工作表名称中的日期
        sheet_name = source_ws.title
        if not sheet_name.startswith("昨日今日半小时成交量"):
            raise ValueError("源工作表名称格式不正确，应以'昨日今日半小时成交量'开头")
        date_str_from_sheet = sheet_name[-8:]
        datetime.strptime(date_str_from_sheet, "%Y%m%d")  # 验证日期格式

        # 计算下一个交易日
        next_date = get_next_trading_day(date_str_from_sheet)
        new_name = f"昨日今日半小时成交量{next_date.strftime('%Y%m%d')}"

        # 检查是否已存在同名工作表
        if new_name in wb.sheetnames:
            raise RuntimeError(f"工作表 {new_name} 已存在，请先删除")

        # 复制工作表
        new_ws = wb.copy_worksheet(source_ws)
        new_ws.title = new_name
        wb.move_sheet(new_ws, offset=-wb.index(new_ws))

        # 将源表 H3:I18 复制到新表 A3:B18
        for row in range(3, 19):
            for src_col, dest_col in zip([8, 9], [1, 2]):
                src_cell = source_ws.cell(row=row, column=src_col)
                dest_cell = new_ws.cell(row=row, column=dest_col)
                dest_cell.value = src_cell.value
                copy_cell_style(src_cell, dest_cell)

        # 插入新行
        new_ws.insert_rows(22)

        # 设置日期
        new_ws['A22'] = next_date
        new_ws['A22'].number_format = 'yyyy/mm/dd'

        # 设置 B22:E22 公式（继承源工作表公式）
        for col in range(2, 6):
            src_formula = source_ws.cell(row=22, column=col).value
            if isinstance(src_formula, str) and src_formula.startswith('='):
                new_ws.cell(row=22, column=col).value = src_formula

        # 设置 B23:E23 引用源工作表的 B22:E22
        prev_sheet_name = source_ws.title
        for col in range(2, 6):
            col_letter = chr(64 + col)
            formula = f"='{prev_sheet_name}'!{col_letter}22"
            new_ws.cell(row=23, column=col).value = formula
            copy_cell_style(source_ws.cell(row=23, column=col), new_ws.cell(row=23, column=col))

        # 清空新表 H3:I18
        for row in range(3, 19):
            for col in range(8, 10):
                new_ws.cell(row=row, column=col).value = None

        return wb
    except Exception as e:
        raise RuntimeError(f"工作表克隆失败: {str(e)}")


def main():
    print("=" * 50)
    print("上证指数数据精准更新与工作表克隆系统")
    print("=" * 50)

    symbol = '000001'  # 上证指数
    date_input = input("请输入交易日(YYYYMMDD): ").strip()

    if not validate_input_date(date_input):
        return

    try:
        if not os.path.exists(TARGET_FILE):
            raise FileNotFoundError(f"目标文件 {TARGET_FILE} 不存在")

        # 加载工作簿，保留公式
        wb = load_workbook(TARGET_FILE, data_only=False)
        source_ws = wb.worksheets[0]

        # 获取并更新数据到第一个工作表
        print(f"\n正在获取 {date_input} 数据...")
        raw_data = fetch_corrected_data(symbol, date_input)
        wb = update_workbook_first_sheet(wb, raw_data)

        # 修改源工作表名称
        new_sheet_name = f"昨日今日半小时成交量{date_input}"
        if source_ws.title != new_sheet_name:
            if new_sheet_name in wb.sheetnames:
                del wb[new_sheet_name]
            source_ws.title = new_sheet_name

        print("\n数据验证：")
        print(f"获取记录数：{len(raw_data)}条")
        print(f"首条原始数据 -> 成交量：{raw_data.iloc[0]['成交量']} 手")
        print(f"           -> 成交额：{raw_data.iloc[0]['成交额']} 元")

        # 克隆工作表
        wb = clone_worksheet(wb, source_ws)

        # 保存文件
        wb.save(TARGET_FILE)
        print(f"\n成功更新文件: {os.path.abspath(TARGET_FILE)}")
        print("新增工作表：", wb.worksheets[0].title)
        os.startfile(TARGET_FILE)

    except Exception as e:
        print(f"\n错误: {str(e)}")
        print("排查建议：")
        print("1. 确认目标文件未被其他程序占用")
        print("2. 检查日期是否为有效交易日")
        print("3. 验证网络连接是否正常")
        print("4. 确保前一交易日工作表存在")
    finally:
        if 'wb' in locals():
            wb.close()
        input("\n按 Enter 键退出...")


if __name__ == "__main__":
    main()