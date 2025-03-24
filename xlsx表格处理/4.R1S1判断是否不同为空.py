import os
import openpyxl
import win32com.client
import time


def export_data_with_filenames():
    """将文件名写入B列，R1/S1写入C/D列"""
    # 路径配置
    source_dir = r'C:\Users\Administrator\Desktop\每日统计改名'
    target_file = r'E:\图片\每笔日交易量价模板.xlsx'
    target_sheet = '30个股R1S1'

    # 定义源数据坐标 (根据实际文件结构调整)
    R1_CELL = 'R1'  # R1值所在单元格
    S1_CELL = 'S1'  # S1值所在单元格

    # 加载目标工作簿
    try:
        wb = openpyxl.load_workbook(target_file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认表

    # 创建/获取目标工作表
    if target_sheet not in wb.sheetnames:
        ws = wb.create_sheet(target_sheet)
        ws['B1'] = '文件名'  # B列标题
        ws['C1'] = 'R1值'
        ws['D1'] = 'S1值'
    else:
        ws = wb[target_sheet]

    # 遍历处理源文件
    for filename in os.listdir(source_dir):
        if not filename.endswith('.xlsx'):
            continue

        try:
            # 读取源文件值
            src_path = os.path.join(source_dir, filename)
            src_wb = openpyxl.load_workbook(src_path, data_only=True)
            src_ws = src_wb.active

            # 获取单元格值，确保不为空
            r1 = src_ws[R1_CELL].value if src_ws[R1_CELL].value is not None else ""
            s1 = src_ws[S1_CELL].value if src_ws[S1_CELL].value is not None else ""

            # 定位写入行（从第二行开始）
            row_num = ws.max_row + 1
            ws[f'B{row_num}'] = filename  # B列写入文件名
            ws[f'C{row_num}'] = r1
            ws[f'D{row_num}'] = s1

            # 放慢程序运行，确保处理完整
            time.sleep(0.5)  # 添加0.5秒延迟

            print(f"成功处理 {filename}：R1={r1}, S1={s1}")

        except Exception as e:
            print(f"处理 {filename} 失败: {str(e)}")
            continue

    # 保存结果
    wb.save(target_file)
    print(f"数据已写入 {target_sheet} 工作表B/C/D列")


def open_excel_and_activate_sheet():
    """打开Excel文件并激活目标工作表"""
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = True  # 可视化界面

    try:
        # 打开工作簿
        workbook = excel.Workbooks.Open(r'E:\图片\每笔日交易量价模板.xlsx')

        # 切换到指定工作表
        worksheet = workbook.Worksheets("30个股R1S1")
        worksheet.Activate()

        print("文件已打开并切换到目标工作表。")

    except Exception as e:
        print(f"操作失败: {e}")
        excel.Quit()


# 执行函数
export_data_with_filenames()
open_excel_and_activate_sheet()
