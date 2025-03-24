import os
from openpyxl import load_workbook
import re  # 用于正则表达式匹配

def process_excel_files(directory):
    # 用于存储已处理的文件名
    processed_files = []
    # 文件计数器
    processed_count = 0

    # 遍历目录下的所有文件
    for filename in os.listdir(directory):
        if not filename.endswith('.xlsx'):  # 只处理.xlsx文件
            continue

        filepath = os.path.join(directory, filename)

        try:
            # 使用openpyxl加载Excel文件，保留所有格式
            wb = load_workbook(filepath)
            ws = wb.active

            # 如果工作表为空，跳过处理
            if ws.max_row <= 1:  # 只有表头或空表
                print(f"文件 {filename} 无数据，跳过处理")
                continue

            # 正则表达式：匹配数字后跟↑或↓，例如 500↑ 或 100↓
            pattern = re.compile(r'^\d+(\.\d+)?[↑↓]$')

            # 找到C列最后一个符合条件（数字+↑或↓）的行
            last_valid_row = None
            for row in range(2, ws.max_row + 1):  # 从第二行开始处理
                c_value = ws[f'C{row}'].value
                if isinstance(c_value, str) and pattern.match(c_value):  # 如果C列值符合条件
                    last_valid_row = row

            # 如果找到了符合条件的行，并且它不是最后一行
            if last_valid_row and last_valid_row < ws.max_row:
                # 删除从最后一个有效行后面的所有行
                ws.delete_rows(last_valid_row + 1, ws.max_row - last_valid_row)

            # 保存修改后的文件
            wb.save(filepath)
            processed_files.append(filename)  # 记录处理成功的文件名
            processed_count += 1  # 计数器加1
            print(f"已成功处理文件: {filename}")

        except Exception as e:
            # 如果处理失败，保留原始文件并打印错误信息
            print(f"处理文件 {filename} 失败: {str(e)}")
            continue

    # 统计并输出处理结果
    print("\n=== 处理统计 ===")
    print(f"共处理了 {processed_count} 个文件")
    if processed_count > 0:
        print("处理的文件列表如下：")
        for file in processed_files:
            print(f"- {file}")
    else:
        print("没有文件被成功处理")

# 指定目录路径
directory = r"C:\Users\Administrator\Desktop\每日统计改名"
process_excel_files(directory)
